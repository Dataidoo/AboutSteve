"""
email_to_sheets.py
==================
Reads tracker submission emails from a Gmail inbox and appends rows to two
separate Excel files — one for the spending tracker (DalaData), one for the
daily tracker.

USAGE
-----
1.  Install dependencies:
        pip install openpyxl

2.  Set up an App Password for the Gmail account:
        Google Account -> Security -> 2-Step Verification -> App passwords
    (regular Gmail passwords no longer work for IMAP)

3.  Fill in the CONFIG section below.

4.  Run:
        python email_to_sheets.py

    Re-run any time. The script tracks which emails it has already processed
    (via UID stored in `processed_uids.json`) so it never double-imports.

LAYOUT EXPECTED
---------------
DalaData ("[BUDGET]") email body — first lines are date + entries, format defined
in the app's send button.

Daily Tracker ("[TRACKER]") email body — starts with "DAILY TRACKER SUBMISSION",
then phase/mood/focus/energy/sleep/redbull/smokes/toggles/raw JSON. The script
parses the RAW JSON block because it's the most reliable shape.
"""

import imaplib
import email
import json
import re
import os
from datetime import datetime
from email.header import decode_header
from openpyxl import Workbook, load_workbook

# ── CONFIG ────────────────────────────────────────────────────────
IMAP_HOST       = 'imap.gmail.com'
IMAP_PORT       = 993
EMAIL_ADDRESS   = 'data.im.storing@gmail.com'
APP_PASSWORD    = 'PUT_YOUR_APP_PASSWORD_HERE'   # 16-char Google App Password, NOT your real password

# Where to write the workbooks (folder must exist)
OUTPUT_FOLDER   = './tracker_output'

SPENDING_XLSX   = os.path.join(OUTPUT_FOLDER, 'spending_tracker.xlsx')
DAILY_XLSX      = os.path.join(OUTPUT_FOLDER, 'habit_tracker.xlsx')

# Subject prefixes the apps send with (keep in sync with the HTML files)
# Spending tracker sends: "DalaData MM/DD/YYYY"
# Daily tracker sends:    "[TRACKER] <localeDate> - <phase>"
SPENDING_SUBJECT_TAG = '[SPENDING]'
DAILY_SUBJECT_TAG    = '[HABIT]'

# UID tracking — prevents reprocessing the same emails
PROCESSED_UIDS_FILE = os.path.join(OUTPUT_FOLDER, 'processed_uids.json')


# ── HELPERS ───────────────────────────────────────────────────────
def decode_subject(raw):
    """Decode RFC-2047 encoded headers (handles emoji and accents)."""
    parts = decode_header(raw or '')
    out = []
    for chunk, enc in parts:
        if isinstance(chunk, bytes):
            out.append(chunk.decode(enc or 'utf-8', errors='replace'))
        else:
            out.append(chunk)
    return ''.join(out)


def get_body_text(msg):
    """Pull the text/plain body out of an email.message.Message."""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                payload = part.get_payload(decode=True)
                if payload:
                    return payload.decode(part.get_content_charset() or 'utf-8', errors='replace')
        return ''
    payload = msg.get_payload(decode=True)
    return payload.decode(msg.get_content_charset() or 'utf-8', errors='replace') if payload else ''


def load_processed_uids():
    if os.path.exists(PROCESSED_UIDS_FILE):
        with open(PROCESSED_UIDS_FILE, 'r') as f:
            return set(json.load(f))
    return set()


def save_processed_uids(uids):
    with open(PROCESSED_UIDS_FILE, 'w') as f:
        json.dump(sorted(uids), f, indent=2)


# ── PARSERS ───────────────────────────────────────────────────────
def parse_spending_email(body, msg_date):
    """
    Returns a tuple: (entry_rows, bills_paid_row).

    The Spending Tracker email body real format (as of current HTML):

        SPENDING ENTRY — 06/20/2026
        ────────────────────────────────
        2026-06-20 | Food | 23.47 | Tim's
        2026-06-20 | Phone | 80.00 | bill
        ────────────────────────────────
        BILLS PAID: Rent, Phone

    Each pipe-line becomes an entry row. The BILLS PAID line is captured
    separately so the workbook can have a 'Bills Paid' sheet alongside 'Entries'.
    """
    entry_rows = []
    # Tolerate any whitespace around pipes
    entry_pattern = re.compile(
        r'^\s*(\d{4}-\d{2}-\d{2})\s*\|\s*([^|]+?)\s*\|\s*([\d.]+)\s*\|\s*(.*?)\s*$',
        re.MULTILINE,
    )
    for m in entry_pattern.finditer(body):
        date, cat, amt, note = m.groups()
        entry_rows.append({
            'date':     date,
            'category': cat.strip(),
            'amount':   float(amt),
            'note':     note.strip(),
            'email_ts': msg_date,
        })

    # Bills paid line
    bills_row = None
    bm = re.search(r'BILLS PAID:\s*(.+)', body)
    if bm:
        paid_list = bm.group(1).strip()
        bills_row = {
            'email_ts':   msg_date,
            'bills_paid': '' if paid_list.lower() == 'none' else paid_list,
        }

    return entry_rows, bills_row


def parse_daily_email(body, msg_date):
    """
    Returns a SINGLE row-dict (one per email = one check-in).

    The daily tracker email embeds a JSON dump at the bottom under "RAW JSON".
    We parse that because it's the most stable shape.
    """
    # Find the JSON block at the end of the email
    m = re.search(r'RAW JSON\s*(\{.*?\})\s*\Z', body, re.DOTALL)
    if not m:
        return None
    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError:
        return None
    data['email_ts'] = msg_date
    return data


# ── EXCEL WRITERS ─────────────────────────────────────────────────
SPENDING_HEADERS    = ['date', 'category', 'amount', 'note', 'email_ts']
BILLS_PAID_HEADERS  = ['email_ts', 'bills_paid']
DAILY_HEADERS       = ['timestamp', 'date', 'time', 'phase', 'mood', 'focus', 'energy',
                       'sleep_hrs', 'redbull', 'smokes', 'toggles', 'email_ts']


def append_rows(xlsx_path, sheet_name, headers, rows):
    """Open or create the workbook, ensure headers, append rows."""
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)

    for row in rows:
        ws.append([row.get(h, '') for h in headers])

    wb.save(xlsx_path)


# ── MAIN ──────────────────────────────────────────────────────────
def main():
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    processed = load_processed_uids()

    print(f'Connecting to {IMAP_HOST}…')
    M = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    M.login(EMAIL_ADDRESS, APP_PASSWORD)
    M.select('INBOX')

    # Search for any tracker emails. ALL is broad; if your inbox is huge, narrow
    # by date — e.g. ('SINCE', '01-Jun-2026')
    status, data = M.search(None, 'ALL')
    if status != 'OK':
        print('Search failed:', status)
        return

    all_uids = data[0].split()
    new_uids = [u for u in all_uids if u.decode() not in processed]
    print(f'{len(new_uids)} new emails to scan (of {len(all_uids)} total).')

    spending_rows = []
    bills_paid_rows = []
    daily_rows = []
    seen_this_run = set()

    for uid in new_uids:
        uid_str = uid.decode()
        status, msg_data = M.fetch(uid, '(RFC822)')
        if status != 'OK':
            continue
        msg = email.message_from_bytes(msg_data[0][1])
        subject = decode_subject(msg.get('Subject', ''))
        msg_date = msg.get('Date', '')
        body = get_body_text(msg)

        if SPENDING_SUBJECT_TAG in subject:
            rows, bills_row = parse_spending_email(body, msg_date)
            if rows or bills_row:
                spending_rows.extend(rows)
                if bills_row:
                    bills_paid_rows.append(bills_row)
                seen_this_run.add(uid_str)
                print(f'  spending: {len(rows)} entries from "{subject}"')
        elif DAILY_SUBJECT_TAG in subject:
            row = parse_daily_email(body, msg_date)
            if row:
                daily_rows.append(row)
                seen_this_run.add(uid_str)
                print(f'  daily:    1 entry from "{subject}"')
        # everything else is ignored (newsletters, replies, etc.)

    M.logout()

    if spending_rows:
        append_rows(SPENDING_XLSX, 'Entries', SPENDING_HEADERS, spending_rows)
        print(f'Wrote {len(spending_rows)} rows -> {SPENDING_XLSX} [Entries]')
    if bills_paid_rows:
        append_rows(SPENDING_XLSX, 'Bills Paid', BILLS_PAID_HEADERS, bills_paid_rows)
        print(f'Wrote {len(bills_paid_rows)} rows -> {SPENDING_XLSX} [Bills Paid]')
    if daily_rows:
        append_rows(DAILY_XLSX, 'Entries', DAILY_HEADERS, daily_rows)
        print(f'Wrote {len(daily_rows)} rows -> {DAILY_XLSX}')

    # Mark processed AFTER successful write so a crash doesn't lose anything
    processed.update(seen_this_run)
    save_processed_uids(processed)
    print('Done.')


if __name__ == '__main__':
    main()
